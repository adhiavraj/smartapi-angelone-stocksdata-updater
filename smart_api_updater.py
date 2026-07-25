import os
import re
import math
import requests
import pyotp
import openpyxl
from datetime import datetime, date, timedelta
from logzero import logger

try:
    from SmartApi import SmartConnect
except Exception as e:
    SmartConnect = None
    logger.warning(f"SmartApi import failed: {e}")

# Default Scrip Master URL
SCRIP_MASTER_URL = "https://margincalculator.angelbroking.com/OpenAPI_File/files/OpenAPIScripMaster.json"

# Stock symbol mapping for SmartAPI
BANK_MAPPING = {
    'HDFC Bank': {'symbol': 'HDFCBANK-EQ', 'token': '1333', 'exch': 'NSE'},
    'ICICI Bank': {'symbol': 'ICICIBANK-EQ', 'token': '4963', 'exch': 'NSE'},
    'Axis Bank': {'symbol': 'AXISBANK-EQ', 'token': '5900', 'exch': 'NSE'},
    'SBI Bank': {'symbol': 'SBIN-EQ', 'token': '3045', 'exch': 'NSE'},
    'Kotak Bank': {'symbol': 'KOTAKBANK-EQ', 'token': '1922', 'exch': 'NSE'},
    'BOB': {'symbol': 'BANKBARODA-EQ', 'token': '4668', 'exch': 'NSE'},
    'Union Bank': {'symbol': 'UNIONBANK-EQ', 'token': '10753', 'exch': 'NSE'},
    'PNB': {'symbol': 'PNB-EQ', 'token': '10666', 'exch': 'NSE'},
    'Canara Bank': {'symbol': 'CANBK-EQ', 'token': '10794', 'exch': 'NSE'},
    'Federal Bank': {'symbol': 'FEDERALBNK-EQ', 'token': '1023', 'exch': 'NSE'},
    'AU Bank': {'symbol': 'AUBANK-EQ', 'token': '21238', 'exch': 'NSE'},
    'Yes Bank': {'symbol': 'YESBANK-EQ', 'token': '11915', 'exch': 'NSE'},
    'Indusind Bank': {'symbol': 'INDUSINDBK-EQ', 'token': '5258', 'exch': 'NSE'},
    'IDFC Bank': {'symbol': 'IDFCFIRSTB-EQ', 'token': '11184', 'exch': 'NSE'},
}

NIFTY_BANK_TOKEN = '99926009'  # Nifty Bank Index Token


def get_public_ip():
    """Fetch user's current public IP address for SmartAPI Primary Static IP registration."""
    services = [
        'https://api.ipify.org',
        'https://ifconfig.me/ip',
        'https://icanhazip.com'
    ]
    for url in services:
        try:
            r = requests.get(url, timeout=3)
            if r.status_code == 200:
                return r.text.strip()
        except Exception:
            continue
    return "127.0.0.1"


class SmartAPIClient:
    def __init__(self, api_key, client_id, password, totp_secret=None):
        self.api_key = api_key
        self.client_id = client_id
        self.password = password
        self.totp_secret = totp_secret
        self.smart_api = None
        self.session_data = None

    def connect(self, manual_totp=None):
        """Authenticate with SmartAPI using TOTP."""
        if not SmartConnect:
            raise RuntimeError("smartapi-python library is not installed in this Python environment. Please run: python -m pip install smartapi-python setuptools")

        totp = manual_totp
        if not totp and self.totp_secret:
            try:
                clean_secret = self.totp_secret.replace(" ", "").upper()
                totp = pyotp.TOTP(clean_secret).now()
            except Exception as e:
                raise ValueError("Invalid TOTP Secret Key format. Please verify your 32-character TOTP secret key from Angel One setup.") from e

        if not totp:
            raise ValueError("TOTP code or TOTP Secret is required for login.")

        try:
            self.smart_api = SmartConnect(api_key=self.api_key)
            data = self.smart_api.generateSession(self.client_id, self.password, totp)
        except Exception as e:
            err_str = str(e)
            if "Couldn't parse the JSON response" in err_str or "b''" in err_str or "JSONDecodeError" in err_str:
                pub_ip = get_public_ip()
                raise RuntimeError(
                    f"SmartAPI Connection Failed: Angel One server returned empty response (b''). "
                    f"Please verify: 1) Your Primary Static IP on smartapi.angelone.in is set to '{pub_ip}', "
                    f"2) Your SmartAPI Key and Client ID are correct, "
                    f"3) Your TOTP Secret Key/Code is valid."
                ) from e
            raise RuntimeError(f"SmartAPI Authentication Error: {err_str}") from e

        if not data or not isinstance(data, dict) or not data.get('status'):
            msg = data.get('message', 'Authentication failed.') if isinstance(data, dict) else 'No response from SmartAPI'
            raise RuntimeError(f"SmartAPI Login Failed: {msg}")

        self.session_data = data
        return data

    def fetch_ltp(self, exchange, tradingsymbol, symboltoken):
        """Fetch current LTP (Last Traded Price) for a symbol."""
        if not self.smart_api:
            raise RuntimeError("SmartAPI client is not connected.")

        try:
            resp = self.smart_api.ltpData(exchange, tradingsymbol, symboltoken)
            if resp and isinstance(resp, dict) and resp.get('status') and resp.get('data'):
                return float(resp['data']['ltp'])
        except Exception as e:
            logger.warning(f"Error fetching LTP for {tradingsymbol}: {e}")
        return None

    def fetch_historical_daily(self, tradingsymbol, symboltoken, from_date, to_date):
        """
        Fetch daily candle data from SmartAPI between from_date and to_date.
        Dates formatted as 'YYYY-MM-DD 09:15'
        Returns dict of {datetime.date: close_price}
        """
        if not self.smart_api:
            raise RuntimeError("SmartAPI client is not connected.")

        param = {
            "servicetype": "kt_service",
            "tradingsymbol": tradingsymbol,
            "symboltoken": str(symboltoken),
            "interval": "ONE_DAY",
            "fromdate": from_date,
            "todate": to_date
        }
        candles = {}
        try:
            resp = self.smart_api.getCandleData(param)
            if resp and isinstance(resp, dict) and resp.get('status') and resp.get('data'):
                for row in resp['data']:
                    ts_str = row[0][:10]
                    d = datetime.strptime(ts_str, "%Y-%m-%d").date()
                    close_px = float(row[4])
                    candles[d] = close_px
        except Exception as e:
            logger.warning(f"Error fetching candle data for {tradingsymbol}: {e}")
        return candles


def load_scrip_tokens():
    """Dynamically download scrip master and return latest tokens dictionary."""
    tokens = {k: v['token'] for k, v in BANK_MAPPING.items()}
    tokens['Nifty Bank'] = NIFTY_BANK_TOKEN
    try:
        r = requests.get(SCRIP_MASTER_URL, timeout=10)
        if r.status_code == 200:
            data = r.json()
            targets = {v['symbol']: k for k, v in BANK_MAPPING.items()}
            for item in data:
                sym = item.get('symbol')
                name = item.get('name')
                exch = item.get('exch_seg')
                if sym in targets and exch == 'NSE':
                    bank_name = targets[sym]
                    tokens[bank_name] = str(item.get('token'))
                if name == 'Nifty Bank' and exch in ['NSE', 'INDICES']:
                    tokens['Nifty Bank'] = str(item.get('token'))
    except Exception as e:
        logger.warning(f"Could not load live scrip master, using built-in tokens: {e}")
    return tokens


def get_excel_summary(excel_path):
    """
    Parse Excel file and return current summary metrics for all bank sheets:
    CMP, Beta, VAR AT 90, VAR IN RS, Last Date, Last Close
    """
    wb_data = openpyxl.load_workbook(excel_path, data_only=True)
    summary = []

    for sheet_name in wb_data.sheetnames:
        if sheet_name == 'Trade Journal':
            continue
        ws = wb_data[sheet_name]

        # Find last populated row in column A
        last_row = 3
        for r in range(4, ws.max_row + 1):
            if ws.cell(row=r, column=1).value is not None:
                last_row = r

        last_date = ws.cell(row=last_row, column=1).value
        last_close = ws.cell(row=last_row, column=2).value
        cmp_val = ws.cell(row=2, column=10).value
        beta_val = ws.cell(row=4, column=10).value
        var90_val = ws.cell(row=6, column=10).value
        var_rs_val = ws.cell(row=8, column=10).value

        # Format date string
        date_str = last_date.strftime("%Y-%m-%d") if isinstance(last_date, (datetime, date)) else str(last_date)

        summary.append({
            'Sheet': sheet_name,
            'Symbol': BANK_MAPPING.get(sheet_name, {}).get('symbol', '-'),
            'Last Date': date_str,
            'Last Close': round(last_close, 2) if last_close is not None else "-",
            'CMP': round(cmp_val, 2) if cmp_val is not None else "-",
            'Beta': round(beta_val, 4) if beta_val is not None else "-",
            'VAR 90%': f"{round(var90_val * 100, 2)}%" if var90_val is not None else "-",
            'VAR (Rs)': round(var_rs_val, 2) if var_rs_val is not None else "-",
            'Last Row': last_row
        })

    return summary


def update_excel_file(excel_path, market_data_by_date, output_path=None):
    """
    Update Excel sheets with new market data.
    market_data_by_date: dict of { date_obj: { 'Nifty Bank': float, 'HDFC Bank': float, ... } }
    Adds new rows with all proper formulas for each date.
    """
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    updated_info = []

    # Sort dates chronologically
    sorted_dates = sorted(market_data_by_date.keys())

    for d in sorted_dates:
        day_data = market_data_by_date[d]
        nifty_close = day_data.get('Nifty Bank')

        for sheet_name in wb.sheetnames:
            if sheet_name == 'Trade Journal':
                continue

            ws = wb[sheet_name]
            stock_close = day_data.get(sheet_name)

            if stock_close is None or nifty_close is None:
                continue

            # Find current last row
            last_row = 3
            for r in range(4, ws.max_row + 1):
                if ws.cell(row=r, column=1).value is not None:
                    last_row = r

            # Check if this date already exists in the last row
            last_date_val = ws.cell(row=last_row, column=1).value
            if isinstance(last_date_val, datetime):
                last_date_obj = last_date_val.date()
            elif isinstance(last_date_val, date):
                last_date_obj = last_date_val
            else:
                last_date_obj = None

            if last_date_obj == d:
                # Update existing row close prices
                ws.cell(row=last_row, column=2, value=stock_close)
                ws.cell(row=last_row, column=6, value=nifty_close)
                updated_info.append(f"Updated row {last_row} in [{sheet_name}] for {d}")
            elif last_date_obj is None or d > last_date_obj:
                # Append new row
                new_row = last_row + 1
                dt_obj = datetime.combine(d, datetime.min.time()) if isinstance(d, date) and not isinstance(d, datetime) else d

                # Col A: Date
                ws.cell(row=new_row, column=1, value=dt_obj)
                # Col B: Bank Close Price
                ws.cell(row=new_row, column=2, value=stock_close)
                # Col C: Price Return =B{new_row}/B{last_row}-1
                ws.cell(row=new_row, column=3, value=f"=B{new_row}/B{last_row}-1")

                # Col E: Nifty Bank Date
                ws.cell(row=new_row, column=5, value=dt_obj)
                # Col F: Nifty Bank Close Price
                ws.cell(row=new_row, column=6, value=nifty_close)

                # Col G: Price Return for Nifty Bank
                if sheet_name == 'HDFC Bank':
                    ws.cell(row=new_row, column=7, value=f"=F{new_row}/F{last_row}-1")
                else:
                    ws.cell(row=new_row, column=7, value=f"='HDFC Bank'!G{new_row}")

                # Col L: Daily Return =VALUE(G{new_row})
                ws.cell(row=new_row, column=12, value=f"=VALUE(G{new_row})")
                # Col M: Simulator =L{new_row}*$J$4
                ws.cell(row=new_row, column=13, value=f"=L{new_row}*$J$4")

                updated_info.append(f"Appended row {new_row} in [{sheet_name}] for date {d}")

    save_target = output_path if output_path else excel_path
    wb.save(save_target)
    wb.close()
    return updated_info
